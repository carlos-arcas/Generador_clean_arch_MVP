"""Blueprint para exportación tabular en Excel (openpyxl) por entidad."""

from __future__ import annotations

from dataclasses import dataclass
import textwrap

from aplicacion.puertos.blueprint import Blueprint
from dominio.modelos import ArchivoGenerado, EspecificacionClase, EspecificacionProyecto, ErrorValidacionDominio, PlanGeneracion


@dataclass(frozen=True)
class NombresClase:
    nombre_clase: str
    nombre_snake: str
    nombre_plural: str


class ExportExcelBlueprint(Blueprint):
    def nombre(self) -> str:
        return "export_excel"

    def version(self) -> str:
        return "1.0.0"

    def validar(self, especificacion: EspecificacionProyecto) -> None:
        especificacion.validar()
        if not especificacion.clases:
            raise ErrorValidacionDominio(
                "El blueprint export_excel requiere al menos una clase en la especificación."
            )

    def generar_plan(self, especificacion: EspecificacionProyecto) -> PlanGeneracion:
        self.validar(especificacion)

        archivos: list[ArchivoGenerado] = [
            ArchivoGenerado(
                "aplicacion/puertos/exportadores/exportador_tabular_excel.py",
                self._contenido_puerto_exportador(),
            ),
            ArchivoGenerado(
                "infraestructura/informes/excel/exportador_excel_openpyxl.py",
                self._contenido_exportador_excel(),
            ),
            ArchivoGenerado(
                "tests/infraestructura/test_export_excel.py",
                self._contenido_test_exportador_excel(),
            ),
        ]

        for clase in especificacion.clases:
            nombres = self._construir_nombres(clase)
            archivos.append(
                ArchivoGenerado(
                    f"aplicacion/casos_uso/informes/generar_informe_{nombres.nombre_plural}_excel.py",
                    self._contenido_caso_uso_informe(clase, nombres),
                )
            )

        plan = PlanGeneracion(archivos=archivos)
        plan.validar_sin_conflictos()
        return plan

    def _construir_nombres(self, clase: EspecificacionClase) -> NombresClase:
        nombre_snake = self._pascal_a_snake(clase.nombre)
        return NombresClase(
            nombre_clase=clase.nombre,
            nombre_snake=nombre_snake,
            nombre_plural=self._pluralizar(nombre_snake),
        )

    def _pascal_a_snake(self, texto: str) -> str:
        caracteres: list[str] = []
        for indice, caracter in enumerate(texto):
            if caracter.isupper() and indice > 0:
                caracteres.append("_")
            caracteres.append(caracter.lower())
        return "".join(caracteres)

    def _pluralizar(self, nombre: str) -> str:
        if nombre.endswith(("a", "e", "i", "o", "u")):
            return f"{nombre}s"
        if nombre.endswith("s"):
            return nombre
        return f"{nombre}es"

    def _contenido_puerto_exportador(self) -> str:
        return textwrap.dedent(
            '''\
            """Puerto de exportación tabular para Excel."""

            from __future__ import annotations

            from abc import ABC, abstractmethod


            class ExportadorTabularExcel(ABC):
                @abstractmethod
                def exportar(self, ruta: str, encabezados: list[str], filas: list[list[str]]) -> None:
                    """Exporta un informe tabular hacia una hoja Excel."""
            '''
        )

    def _contenido_caso_uso_informe(self, clase: EspecificacionClase, nombres: NombresClase) -> str:
        columnas = ["id", *[atributo.nombre for atributo in clase.atributos if atributo.nombre != "id"]]
        return textwrap.dedent(
            f'''\
            """Caso de uso para generar informe Excel de {nombres.nombre_plural}."""

            from __future__ import annotations

            import logging

            from aplicacion.puertos.exportadores.exportador_tabular_excel import ExportadorTabularExcel
            from aplicacion.puertos.repositorio_{nombres.nombre_snake} import Repositorio{nombres.nombre_clase}

            LOGGER = logging.getLogger(__name__)


            class GenerarInforme{nombres.nombre_clase}Excel:
                def __init__(
                    self,
                    repositorio: Repositorio{nombres.nombre_clase},
                    exportador: ExportadorTabularExcel,
                ) -> None:
                    self._repositorio = repositorio
                    self._exportador = exportador
                    self._encabezados = {columnas!r}

                def ejecutar(self, ruta_salida: str) -> str:
                    entidades = self._repositorio.listar()
                    filas = [
                        [str(getattr(entidad, columna, "")) for columna in self._encabezados]
                        for entidad in entidades
                    ]
                    self._exportador.exportar(ruta=ruta_salida, encabezados=self._encabezados, filas=filas)
                    LOGGER.info("Informe Excel generado en %s", ruta_salida)
                    return ruta_salida
            '''
        )

    def _contenido_exportador_excel(self) -> str:
        return textwrap.dedent(
            '''\
            """Exportador tabular Excel basado en openpyxl."""

            from __future__ import annotations

            from pathlib import Path

            from openpyxl import Workbook

            from aplicacion.puertos.exportadores.exportador_tabular_excel import ExportadorTabularExcel


            class ExportadorExcelOpenpyxl(ExportadorTabularExcel):
                def exportar(self, ruta: str, encabezados: list[str], filas: list[list[str]]) -> None:
                    destino = Path(ruta)
                    destino.parent.mkdir(parents=True, exist_ok=True)

                    libro = Workbook()
                    hoja = libro.active
                    hoja.title = "Informe"
                    hoja.append(encabezados)
                    for fila in filas:
                        hoja.append(fila)
                    libro.save(destino)
            '''
        )

    def _contenido_test_exportador_excel(self) -> str:
        return textwrap.dedent(
            '''\
            from pathlib import Path

            from openpyxl import load_workbook

            from infraestructura.informes.excel.exportador_excel_openpyxl import ExportadorExcelOpenpyxl


            def test_exportador_excel_generar_hoja_con_encabezados(tmp_path: Path) -> None:
                ruta = tmp_path / "clientes.xlsx"

                ExportadorExcelOpenpyxl().exportar(
                    ruta=str(ruta),
                    encabezados=["id", "nombre"],
                    filas=[["1", "Ana"]],
                )

                assert ruta.exists()
                libro = load_workbook(ruta)
                hoja = libro.active
                assert hoja["A1"].value == "id"
                assert hoja["B1"].value == "nombre"
                assert hoja["A2"].value == "1"
                assert hoja["B2"].value == "Ana"
            '''
        )
